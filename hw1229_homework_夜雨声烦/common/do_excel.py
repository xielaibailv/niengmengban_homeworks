#！/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2018/12/29 22:00
# @Author   :Yosef
# E-mail    :wurz529@foxmail.com
# File      :do_excel.py
# Software  :PyCharm Community Edition
import openpyxl
class DoExcel:

    def __init__(self,filepath,sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name

    def read_data(self):
        wb = openpyxl.load_workbook(self.filepath)
        sh = wb[self.sheet_name]
        print("开始读取{}中的{}数据".format(self.filepath, self.sheet_name))

        col_max = sh.max_column
        testdata_key = []
        for i in range(1, col_max + 1):
            testdata_key.append(sh.cell(1, i).value)

        testdatas = []
        row_max = sh.max_row
        print(row_max,col_max)
        for i in range(2, row_max + 1):
            testdata = {}
            for j in range(1, col_max + 1):
                testdata[testdata_key[j - 1]] = sh.cell(i, j).value
            testdatas.append(testdata)
        print("读取完毕")

        return testdatas

    def write_data(self, row, col, value):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb[self.sheet_name]

        ws.cell(row, col).value = value
        wb.save(self.filepath)

if __name__ == '__main__':
    excel = DoExcel("../conf/test_data.xlsx","add_test_data").read_data()
    print(len(excel))
