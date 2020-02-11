from openpyxl import load_workbook  # 读写


class Cases:
    def __init__(self):
        self.id = None
        self.title = None
        self.a = None
        self.b = None
        self.excepted = None


# 将从excel文件中读取数据和写入数据写成类，方便用例执行调用。
class DoExcel:
    def __init__(self, filename, sheet_name):
        self.filename = filename  # 工作簿
        self.sheet_name = sheet_name   # sheet名

    def read_data(self):   # 获取数据
        # 1.打开excel
        excel = load_workbook(self.filename)
        # 2.定位表单
        sheet = excel[self.sheet_name]
        # # 3.将读取到的数据每行存进一个字典（一个用例），所有用例存进一个列表
        # test_data = []
        # for i in range(2, sheet.max_row + 1):  # 从第二行开始读取
        #     per_data = {}
        #     per_data['id'] = sheet.cell(i, 1).value   # 存用例的序号
        #     per_data['title'] = sheet.cell(i, 2).value  # 存用例的标题
        #     per_data['a'] = sheet.cell(i, 3).value  # 存用例的参数 a
        #     per_data['b'] = sheet.cell(i, 4).value  # 存用例的参数 b
        #     per_data['excepted'] = sheet.cell(i, 5).value  # 存用例的期望值
        #     test_data.append(per_data)
        # return test_data

        # 利用类与对象的思维来写read_data
        cases = []   # 所有用例存进一个列表
        for i in range(2, sheet.max_row + 1):  # 从第二行开始读取
            case = Cases()   # 每一行数据是一个用例，存在Cases()这个对象里面
            case.id = sheet.cell(i, 1).value   # 存用例的序号
            case.title = sheet.cell(i, 2).value  # 存用例的标题
            case.a = sheet.cell(i, 3).value  # 存用例的参数 a
            case.b = sheet.cell(i, 4).value  # 存用例的参数 b
            case.expected = sheet.cell(i, 5).value  # 存用例的期望值
            cases.append(case)
        return cases

    # row:写入的行 ;  column: 写入的列 ;  value: 写入的值
    def write_data(self, row, column, value):   # 回写数据
        # 1.打开excel
        excel = load_workbook(self.filename)
        # 2.定位表单
        sheet = excel[self.sheet_name]
        # 3. 写入数据
        sheet.cell(row, column).value = value
        # 4.保存
        excel.save(self.filename)


if __name__ == '__main__':
    wb = DoExcel('测试数学类的数据.xlsx', 'data_add')
    print(wb.read_data())
