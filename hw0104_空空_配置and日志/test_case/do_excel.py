from openpyxl import load_workbook  # 读写


class Cases:
    def __init__(self):
        self.id = None
        self.title = None
        self.a = None
        self.b = None
        self.excepted = None



# 将从excel文件中读取数据和写入数据写成类，方便用例执行调用。
class DoExcel:
    def __init__(self, filename, sheet_name,button):
        self.filename = filename  # 工作簿
        self.sheet_name = sheet_name   # sheet名
        self.button = button

    def read_data(self):   # 获取数据
        # 1.打开excel
        excel = load_workbook(self.filename)
        # 2.定位表单
        sheet = excel[self.sheet_name]
        # # 3.利用类与对象的思维，将读取到的数据存进对象
        cases = []   # 所有用例存进一个列表

        if self.button == 'all':   # 执行所有用例
            for i in range(2, sheet.max_row + 1):  # 从第二行开始读取
                case = Cases()   # 每一行数据是一个用例，存在Cases()这个对象里面
                case.id = sheet.cell(i, 1).value   # 存用例的序号
                case.title = sheet.cell(i, 2).value  # 存用例的标题
                case.a = sheet.cell(i, 3).value  # 存用例的参数 a
                case.b = sheet.cell(i, 4).value  # 存用例的参数 b
                case.expected = sheet.cell(i, 5).value  # 存用例的期望值
                cases.append(case)
            return cases
        elif type(eval(self.button)) == int:  # 假如配置文件是单个数字，则从头执行到该行的用例
            self.button = int(self.button)
            for i in range(2, self.button+2):  # 从第二行开始读取
                case = Cases()  # 每一行数据是一个用例，存在Cases()这个对象里面
                case.id = sheet.cell(i, 1).value  # 存用例的序号
                case.title = sheet.cell(i, 2).value  # 存用例的标题
                case.a = sheet.cell(i, 3).value  # 存用例的参数 a
                case.b = sheet.cell(i, 4).value  # 存用例的参数 b
                case.expected = sheet.cell(i, 5).value  # 存用例的期望值
                cases.append(case)
            return cases
        elif type(eval(self.button)) == list:   # 如果列表，[a,b]则从a行执行到b行
            self.button = eval(self.button)
            for i in range(self.button[0]+1, self.button[1]+2):
                case = Cases()  # 每一行数据是一个用例，存在Cases()这个对象里面
                case.id = sheet.cell(i, 1).value  # 存用例的序号
                case.title = sheet.cell(i, 2).value  # 存用例的标题
                case.a = sheet.cell(i, 3).value  # 存用例的参数 a
                case.b = sheet.cell(i, 4).value  # 存用例的参数 b
                case.expected = sheet.cell(i, 5).value  # 存用例的期望值
                cases.append(case)
            return cases

    # row:写入的行 ;  column: 写入的列 ;  value: 写入的值
    def write_data(self, row, column, value):   # 回写数据
        # 1.打开excel
        excel = load_workbook(self.filename)
        # 2.定位表单
        sheet = excel[self.sheet_name]
        # 3. 写入数据
        sheet.cell(row, column).value = value
        # 4.保存
        excel.save(self.filename)


if __name__ == '__main__':
    wb = DoExcel('测试数学类的数据.xlsx', 'data_add',1)
    print(wb.read_data())
